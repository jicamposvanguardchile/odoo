import binascii
import certifi
import csv
import tempfile
import urllib3
import xlrd

from odoo import fields,models, _
from datetime import datetime

import base64
import openpyxl
from io import BytesIO
from odoo.exceptions import UserError
from odoo.tools import ustr
import logging

_logger = logging.getLogger(__name__)


class DataImport(models.TransientModel):
    _name = 'data.import'
    _description = 'Import Data wizard'

    file_type = fields.Selection([('csv', 'CSV File'),
                                  ('xls', 'Excel File')],
                                 String='File Type', Required=True,
                                 default='csv',
                                 help='File type of importing file')
    name = fields.Char(string='name')
    file = fields.Binary(String='Upload File', Required=True,
                         help='File to Import')

    def action_import_data(self):
        if self.file:
            if self.file_type == 'csv':
                try:
                    file = base64.b64decode(self.file)
                    file_string = file.decode('utf-8')
                    file_string.split('\n')
                    urllib3.PoolManager(cert_reqs='CERT_REQUIRED',
                                        ca_certs=certifi.where())
                except:
                    raise UserError(_("Please choose the correct file!"))
                for rec in self:
                    file = str(base64.decodebytes(
                        rec.file).decode('utf-8'))
                    reader = csv.reader(file.splitlines())
                    next(reader)
                    last_bom_id = False
                    rec_count = 0
                    try:
                        for col in reader:
                            bom_data = {}
                            bom_line_dict = {}
                            if col[1]:
                                rec_count += 1
                                if rec.product_by == 'name':
                                    product_search = 'name'
                                elif rec.product_by == 'default_code':
                                    product_search = 'default_code'
                                else:
                                    product_search = 'barcode'
                                product = self.env['product.template'].search(
                                    [(product_search, '=', col[1])], limit=1)
                                bom_data.update({'product_tmpl_id': product.id,
                                                 'code': col[0]})
                                if col[2]:
                                    if rec.product_variant_by == 'default_code':
                                        variant_search = 'default_code'
                                    else:
                                        variant_search = 'barcode'
                                    variant = self.env[
                                        'product.product'].search(
                                        [(variant_search, '=', col[2])],
                                        limit=1)
                                    bom_data.update({'product_id': variant.id})
                                    if col[4]:
                                        uom = self.env['uom.uom'].search(
                                            [('name', '=', col[4])], limit=1)
                                        bom_data.update(
                                            {'product_uom_id': uom.id})
                                    else:
                                        bom_data.update(
                                            {'product_uom_id': variant.uom_id})
                                if col[3]:
                                    bom_data.update({'product_qty': col[3]})
                                else:
                                    bom_data.update({'product_qty': 1})

                                if rec.bom_type == 'mtp':
                                    bom_data.update({'type': 'normal'})
                                else:
                                    bom_data.update({'type': 'phantom'})
                                bom_bom = rec.env['mrp.bom'].create(bom_data)
                                last_bom_id = bom_bom.id
                            if col[5]:
                                if rec.product_variant_by == 'default_code':
                                    variant_search = 'default_code'
                                else:
                                    variant_search = 'barcode'
                                variant = self.env['product.product'].search(
                                    [(variant_search, '=', col[5])], limit=1)
                                bom_line_dict.update({'product_id': variant.id,
                                                      'bom_id': last_bom_id
                                                      })
                                if col[6]:
                                    bom_line_dict.update(
                                        {'product_qty': col[6]})
                                else:
                                    bom_line_dict.update({'product_qty': 1})
                                if col[7]:
                                    uom = self.env['uom.uom'].search(
                                        [('name', '=', col[7])], limit=1)
                                    bom_line_dict.update(
                                        {'product_uom_id': uom.id})
                                else:
                                    bom_line_dict.update(
                                        {'product_uom_id': product.uom_id})
                                self.env['mrp.bom.line'].create(bom_line_dict)
                        return rec.success_message(rec_count)
                    except Exception as e:
                        raise UserError(
                            _("Sorry, The CSV file you provided "
                              "does not match our required format" + ustr(
                                e)))

            if self.file_type == 'xls':
                #try:
                #    file_string = tempfile.NamedTemporaryFile(suffix=".xlsx")
                #    file_string.write(binascii.a2b_base64(self.file))
                #    book = xlrd.open_workbook(file_string.name)
                ##    book.sheet_by_index(0)
                #    urllib3.PoolManager(cert_reqs='CERT_REQUIRED',
                #                        ca_certs=certifi.where())
                #except:
                #    raise UserError(_("Please choose the correct file"))
                for rec in self:
#                    try:
                        wb = openpyxl.load_workbook(
                            filename=BytesIO(base64.b64decode(rec.file)),
                            read_only=False)
                        #ws = wb.active
                        ws = wb['Historico']
                        rec_count = 0

                        Group = self.env['process.group']
                        new_group = Group.create({
                                'name': rec.name,
                                'date': datetime.now()
                            })

                        for col in ws.iter_rows(min_row=2, values_only=True):
                            if not col[0]:
                                break
                        #for col in ws.iter_rows(min_row=2, values_only=False):
                            values = {
                                        'codigo_req': col[0] or '',
                                        'numero_de_os': col[1] or '',
                                        'estado_os': col[2] or '',
                                        #'fecha_apertura': col[3],
                                        #'fecha_comprometida': col[4],
                                        #'fecha_test': col[5],
                                        #'fecha_finalizacion': col[6],
                                        'hh_estimadas': col[7] or 0,
                                        'hh_cotizada': col[8] or 0,
                                        'grupo_responsable': col[9] or '',
                                        'categoria': col[10] or '',
                                        'tipo': col[11] or '',
                                        'descripcion': col[12] or '',
                                        'respuesta': col[13] or '',
                                        #'fecha_produccion': col[14],
                                        #'fecha_pendiente': col[15],
                                        #'fecha_test_aprobado': col[16],
                                        #'fecha_ultima_iteracion': col[17],
                                        #'fecha_usuario_compromiso': col[18],
                                        'hh_administracion_jp': col[19] or 0,
                                        'hh_analisis': col[20] or 0,
                                        'hh_construccion': col[21] or 0,
                                        'hh_correcciones': col[22] or 0,
                                        'hh_diseno': col[23] or 0,
                                        'hh_instalacion': col[24] or 0,
                                        'hh_pruebas': col[25] or 0,
                                        #'fecha_ingreso_moebius': col[26],
                                        'analista_resp_sonda': col[27] or '',
                                        'departamento_usuario': col[28] or '',
                                        'agente_resp_apf': col[29] or '',
                                        'tipo_solicitud': col[30] or '',
                                        'prioridad': col[31] or '',
                                        #'fecha_hora_atencion': col[32],
                                        #'fecha_hora_transferencia': col[33],
                                        'area_usuario': col[34] or '',
                                        'reporte_cliente': col[35] or '',
                                    }
                            if col[3]:
                                values['fecha_apertura'] = col[3]
                            if col[4]:
                                values['fecha_comprometida'] = col[4]
                            if col[5]:
                                values['fecha_test'] = col[5]
                            if col[6]:
                                values['fecha_finalizacion'] = col[6]
                            if col[14]:
                                values['fecha_produccion'] = col[14]
                            if col[15]:
                                values['fecha_pendiente'] = col[15]
                            if col[16]:
                                values['fecha_test_aprobado'] = col[16]
                            if col[17]:
                                values['fecha_ultima_iteracion'] = col[17]
                            if col[18]:
                                values['fecha_usuario_compromiso'] = col[18]
                            if col[26]:
                                values['fecha_ingreso_moebius'] = col[26]
                            if col[32]:
                                values['fecha_hora_atencion'] = col[32]
                            if col[33]:
                                values['fecha_hora_transferencia'] = col[33]
                            #_logger.info('>>> 1')
                            #_logger.info('%s'%(values))
                            new_group.process_data_ids = [(0, 0, values)]

                            rec_count += 1
#                        return rec.success_message(rec_count)
#                    except Exception as e:
#                        raise UserError(
#                            _("Sorry, The Excel file you provided"
#                              " does not match our required format" + ustr(
#                                e)))

    def success_message(self, rec_count):
        """function for displaying success message"""
        message_id = self.env['success.message'].create(
            {'message': str(rec_count) + " Records imported successfully"})
        return {
            'name': 'Message',
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'success.message',
            'res_id': message_id.id,
            'target': 'new'
        }
